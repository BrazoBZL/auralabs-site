"""
Aura Labs Pricing Engine — Generator #2
Customer Catalog PDF

Hides all internal data (supplier cost, margins, benchmark prices).
Shows: Compound | Strength | Format | Retail Price only.

Input:  MasterPricing.xlsx
Output: output/Aura-Labs-Customer-Catalog.pdf

Usage:  python generate-customer-catalog.py
"""

import openpyxl
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
XLSX       = SCRIPT_DIR / "MasterPricing.xlsx"
OUTPUT_HTML = SCRIPT_DIR / "output" / "_customer-catalog.html"
OUTPUT_PDF  = SCRIPT_DIR / "output" / "Aura-Labs-Customer-Catalog.pdf"

VERSION = "1.0"

CAT_ORDER = [
    "Metabolic & Weight Management",
    "Recovery & Regeneration",
    "Growth Hormone & Secretagogues",
    "Cognitive & Nootropics",
    "Copper Peptides & Aesthetic",
    "Longevity & Bioregulators",
    "Hormone Support",
    "Reconstitution Supplies",
]


def load_skus():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    skus = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        skus.append(dict(zip(headers, row)))
    return skus


def fmt_price(val):
    if val is None:
        return "—"
    try:
        return f"${float(val):,.0f}"
    except:
        return str(val)


def main():
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found.")
        sys.exit(1)

    skus = load_skus()
    # Only show Website Visible = Yes, Status = Active
    visible = [s for s in skus if s.get("Website Visible") == "Yes"
               and s.get("Status") == "Active"]

    grouped = {cat: [] for cat in CAT_ORDER}
    for s in visible:
        cat = s.get("Category", "")
        if cat in grouped:
            grouped[cat].append(s)

    today = date.today().strftime("%B %Y")
    gen_ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    sections = ""
    for cat, items in grouped.items():
        if not items:
            continue
        rows = ""
        for item in items:
            compound = item.get("Compound", "")
            strength = item.get("Strength", "")
            fmt = item.get("Format", "")
            price = fmt_price(item.get("Aura Retail Price"))
            blend_note = ""
            if "+" in str(strength):
                blend_note = f'<span class="blend-note">{strength}</span>'
                strength = ""
            rows += f"""
      <tr>
        <td class="compound-name">{compound}{blend_note}</td>
        <td class="strength">{strength}</td>
        <td class="format">{fmt}</td>
        <td>{price}</td>
      </tr>"""

        sections += f"""
  <div class="section-title">{cat}</div>
  <table>
    <thead>
      <tr>
        <th>Compound</th><th>Strength</th><th>Format</th>
        <th style="text-align:right">Price</th>
      </tr>
    </thead>
    <tbody>{rows}
    </tbody>
  </table>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Aura Labs — Research Catalog | {today}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=EB+Garamond:wght@400;500;600&family=Inter:wght@400;500;600;700&display=swap');
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:'Inter',sans-serif; background:#fafaf8; color:#1a1a1a; font-size:10px; line-height:1.5; }}
  .page {{ max-width:860px; margin:0 auto; padding:48px 52px; }}
  .header {{ display:flex; justify-content:space-between; align-items:flex-end; border-bottom:1.5px solid #1a1a1a; padding-bottom:16px; margin-bottom:28px; }}
  .brand-name {{ font-family:'EB Garamond',serif; font-size:22px; font-weight:600; letter-spacing:0.08em; text-transform:uppercase; }}
  .brand-sub {{ font-size:8.5px; font-weight:500; letter-spacing:0.2em; text-transform:uppercase; color:#666; margin-top:3px; }}
  .header-meta {{ text-align:right; font-size:8.5px; color:#666; letter-spacing:0.08em; line-height:1.8; }}
  .header-meta strong {{ color:#1a1a1a; }}
  .notice {{ background:#f0ede8; border-left:2px solid #1a1a1a; padding:10px 14px; margin-bottom:24px; font-size:8.5px; color:#444; line-height:1.6; }}
  .section-title {{ font-family:'EB Garamond',serif; font-size:13px; font-weight:600; letter-spacing:0.12em; text-transform:uppercase; border-bottom:0.5px solid #ccc; padding-bottom:6px; margin-bottom:10px; margin-top:22px; }}
  table {{ width:100%; border-collapse:collapse; margin-bottom:4px; }}
  thead tr {{ border-bottom:1px solid #1a1a1a; }}
  thead th {{ font-size:7.5px; font-weight:700; letter-spacing:0.15em; text-transform:uppercase; color:#666; padding:6px 8px 6px 0; text-align:left; }}
  thead th:last-child {{ text-align:right; padding-right:0; }}
  tbody tr {{ border-bottom:0.5px solid #e8e5e0; }}
  tbody tr:last-child {{ border-bottom:none; }}
  tbody td {{ padding:7px 8px 7px 0; font-size:9.5px; vertical-align:middle; }}
  tbody td:last-child {{ text-align:right; padding-right:0; font-weight:600; font-size:10px; }}
  .compound-name {{ font-weight:500; }}
  .strength {{ color:#555; }}
  .format {{ color:#888; font-size:8.5px; }}
  .blend-note {{ font-size:8px; color:#888; display:block; margin-top:1px; }}
  .footer {{ margin-top:36px; border-top:0.5px solid #ccc; padding-top:14px; display:flex; justify-content:space-between; }}
  .footer p {{ font-size:7.5px; color:#888; line-height:1.8; }}
</style>
</head>
<body>
<div class="page">
  <div class="header">
    <div>
      <div class="brand-name">Aura Labs</div>
      <div class="brand-sub">Research Compounds — Catalog</div>
    </div>
    <div class="header-meta">
      <p><strong>Effective:</strong> {today}</p>
      <p>auralabsbio.co</p>
    </div>
  </div>
  <div class="notice">All compounds sold strictly for research and laboratory use only. Not for human consumption. For qualified research professionals. Prices per single vial unless noted.</div>
  {sections}
  <div class="footer">
    <div>
      <p>auralabsbio.co &nbsp;·&nbsp; Research use only &nbsp;·&nbsp; Not for human consumption</p>
      <p>Prices subject to change without notice</p>
    </div>
    <div style="text-align:right">
      <p>Aura Labs &nbsp;·&nbsp; {today}</p>
    </div>
  </div>
</div>
</body>
</html>"""

    OUTPUT_HTML.write_text(html, encoding="utf-8")

    result = subprocess.run(
        ["wkhtmltopdf", "--page-size", "Letter",
         "--margin-top", "0", "--margin-bottom", "0",
         "--margin-left", "0", "--margin-right", "0",
         "--enable-local-file-access",
         str(OUTPUT_HTML), str(OUTPUT_PDF)],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print("wkhtmltopdf error:", result.stderr[-300:])
        sys.exit(1)

    size_kb = OUTPUT_PDF.stat().st_size // 1024
    print(f"✓ Customer Catalog: {OUTPUT_PDF.name} ({size_kb} KB) — {len(visible)} SKUs")


if __name__ == "__main__":
    main()
