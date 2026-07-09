"""
Aura Labs Pricing Engine — Generator #1
Internal Price List PDF

Input:  MasterPricing.xlsx
Output: output/Aura-Labs-Price-List.pdf

Usage:  python generate-price-list.py
"""

import openpyxl
import subprocess
import sys
import os
from datetime import date, datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
XLSX = SCRIPT_DIR / "MasterPricing.xlsx"
TEMPLATE = SCRIPT_DIR / "templates" / "price-list.html"
OUTPUT_HTML = SCRIPT_DIR / "output" / "_price-list.html"
OUTPUT_PDF  = SCRIPT_DIR / "output" / "Aura-Labs-Price-List.pdf"

VERSION = "1.0"
PRICING_STANDARD = "AURA LABS PRICING STANDARD v1.0"


def load_skus():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    skus = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        d = dict(zip(headers, row))
        skus.append(d)
    return skus


def group_by_category(skus):
    """Return ordered dict: category → list of active SKUs."""
    CAT_ORDER = [
        "Metabolic & Weight Management",
        "Recovery & Regeneration",
        "Growth Hormone & Secretagogues",
        "Cognitive & Nootropics",
        "Copper Peptides & Aesthetic",
        "Longevity & Bioregulators",
        "Hormone Support",
        "Reconstitution Supplies",
    ]
    grouped = {cat: [] for cat in CAT_ORDER}
    for sku in skus:
        cat = sku.get("Category", "")
        status = sku.get("Status", "")
        if status == "Active" and cat in grouped:
            grouped[cat].append(sku)
    return grouped


def fmt_price(val):
    if val is None:
        return "—"
    try:
        return f"${float(val):,.0f}"
    except:
        return str(val)


def generate_html(skus):
    grouped = group_by_category(skus)
    today = date.today().strftime("%B %Y")
    gen_ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Build category HTML blocks
    sections_html = ""
    for cat, items in grouped.items():
        if not items:
            continue
        rows_html = ""
        for item in items:
            compound = item.get("Compound", "")
            strength = item.get("Strength", "")
            fmt = item.get("Format", "")
            price = fmt_price(item.get("Aura Retail Price"))

            # Blend note: strength with "+" implies a blend
            blend_note = ""
            if "+" in str(strength):
                blend_note = f'<span class="blend-note">{strength}</span>'
                strength = ""

            rows_html += f"""
      <tr>
        <td class="compound-name">{compound}{blend_note}</td>
        <td class="strength">{strength}</td>
        <td class="format">{fmt}</td>
        <td>{price}</td>
      </tr>"""

        sections_html += f"""
  <div class="section-title">{cat}</div>
  <table>
    <thead>
      <tr>
        <th>Compound</th>
        <th>Strength</th>
        <th>Format</th>
        <th style="text-align:right">Price</th>
      </tr>
    </thead>
    <tbody>{rows_html}
    </tbody>
  </table>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Aura Labs — Research Compound Price List | {today}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=EB+Garamond:wght@400;500;600&family=Inter:wght@400;500;600;700&display=swap');
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: 'Inter', sans-serif; background: #fafaf8; color: #1a1a1a; font-size: 10px; line-height: 1.5; }}
  .page {{ width: 100%; max-width: 860px; margin: 0 auto; padding: 48px 52px; }}
  .header {{ display: flex; justify-content: space-between; align-items: flex-end; border-bottom: 1.5px solid #1a1a1a; padding-bottom: 16px; margin-bottom: 28px; }}
  .brand-name {{ font-family: 'EB Garamond', serif; font-size: 22px; font-weight: 600; letter-spacing: 0.08em; text-transform: uppercase; color: #1a1a1a; }}
  .brand-sub {{ font-family: 'Inter', sans-serif; font-size: 8.5px; font-weight: 500; letter-spacing: 0.2em; text-transform: uppercase; color: #666; margin-top: 3px; }}
  .header-meta {{ text-align: right; }}
  .header-meta p {{ font-size: 8.5px; color: #666; letter-spacing: 0.08em; line-height: 1.8; }}
  .header-meta strong {{ color: #1a1a1a; font-weight: 600; }}
  .notice {{ background: #f0ede8; border-left: 2px solid #1a1a1a; padding: 10px 14px; margin-bottom: 24px; font-size: 8.5px; color: #444; line-height: 1.6; letter-spacing: 0.03em; }}
  .section-title {{ font-family: 'EB Garamond', serif; font-size: 13px; font-weight: 600; letter-spacing: 0.12em; text-transform: uppercase; color: #1a1a1a; border-bottom: 0.5px solid #ccc; padding-bottom: 6px; margin-bottom: 10px; margin-top: 22px; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 4px; }}
  thead tr {{ border-bottom: 1px solid #1a1a1a; }}
  thead th {{ font-size: 7.5px; font-weight: 700; letter-spacing: 0.15em; text-transform: uppercase; color: #666; padding: 6px 8px 6px 0; text-align: left; }}
  thead th:last-child {{ text-align: right; padding-right: 0; }}
  tbody tr {{ border-bottom: 0.5px solid #e8e5e0; }}
  tbody tr:last-child {{ border-bottom: none; }}
  tbody td {{ padding: 7px 8px 7px 0; font-size: 9.5px; color: #1a1a1a; vertical-align: middle; }}
  tbody td:last-child {{ text-align: right; padding-right: 0; font-weight: 600; font-size: 10px; }}
  .compound-name {{ font-weight: 500; }}
  .strength {{ color: #555; font-weight: 400; }}
  .format {{ color: #888; font-size: 8.5px; font-weight: 400; }}
  .blend-note {{ font-size: 8px; color: #888; display: block; margin-top: 1px; }}
  .footer {{ margin-top: 36px; border-top: 0.5px solid #ccc; padding-top: 14px; display: flex; justify-content: space-between; align-items: flex-start; }}
  .footer p {{ font-size: 7.5px; color: #888; line-height: 1.8; letter-spacing: 0.04em; }}
</style>
</head>
<body>
<div class="page">
  <div class="header">
    <div class="brand">
      <div class="brand-name">Aura Labs</div>
      <div class="brand-sub">Research Compounds — Official Price List</div>
    </div>
    <div class="header-meta">
      <p><strong>Effective:</strong> {today}</p>
      <p><strong>Version:</strong> {VERSION}</p>
      <p>auralabsbio.co</p>
    </div>
  </div>
  <div class="notice">
    All compounds are sold strictly for research and laboratory use only. Not for human consumption. For qualified research professionals only. Prices are per single vial unless otherwise noted.
  </div>
  {sections_html}
  <div class="footer">
    <div>
      <p>auralabsbio.co &nbsp;·&nbsp; All compounds for research use only</p>
      <p>Not for human consumption &nbsp;·&nbsp; Qualified researchers only</p>
      <p>Prices subject to change &nbsp;·&nbsp; {PRICING_STANDARD}</p>
    </div>
    <div style="text-align:right">
      <p>Aura Labs &nbsp;·&nbsp; {today}</p>
      <p>Generated: {gen_ts}</p>
    </div>
  </div>
</div>
</body>
</html>"""
    return html


def main():
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found. Run MasterPricing_builder.py first.")
        sys.exit(1)

    print("Reading MasterPricing.xlsx...")
    skus = load_skus()
    active = [s for s in skus if s.get("Status") == "Active"]
    print(f"  {len(active)} active SKUs across {len(set(s['Category'] for s in active))} categories")

    print("Generating HTML...")
    html = generate_html(skus)
    OUTPUT_HTML.write_text(html, encoding="utf-8")

    print("Rendering PDF...")
    result = subprocess.run(
        ["wkhtmltopdf",
         "--page-size", "Letter",
         "--margin-top", "0",
         "--margin-bottom", "0",
         "--margin-left", "0",
         "--margin-right", "0",
         "--enable-local-file-access",
         str(OUTPUT_HTML),
         str(OUTPUT_PDF)],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print("wkhtmltopdf stderr:", result.stderr[-500:])
        sys.exit(1)

    size_kb = OUTPUT_PDF.stat().st_size // 1024
    print(f"✓ PDF generated: {OUTPUT_PDF.name} ({size_kb} KB)")
    return OUTPUT_PDF


if __name__ == "__main__":
    main()
