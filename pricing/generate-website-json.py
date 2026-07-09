"""
Aura Labs Pricing Engine — Generator #4
Website JSON

Exports clean pricing.json for the website.
Only includes: Website Visible = Yes, Status = Active

Output fields: category, compound, strength, format, price, slug,
               sku, status, featured, tier

Input:  MasterPricing.xlsx
Output: output/pricing.json
        output/pricing-flat.json  (simple array, no grouping)

Usage:  python generate-website-json.py
"""

import openpyxl
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

SCRIPT_DIR   = Path(__file__).parent
XLSX         = SCRIPT_DIR / "MasterPricing.xlsx"
OUTPUT_JSON  = SCRIPT_DIR / "output" / "pricing.json"
OUTPUT_FLAT  = SCRIPT_DIR / "output" / "pricing-flat.json"

CAT_ORDER = [
    "Metabolic & Weight Management",
    "Recovery & Regeneration",
    "Growth Hormone & Secretagogues",
    "Cognitive & Nootropics",
    "Copper Peptides & Aesthetic",
    "Longevity & Bioregulators",
    "Hormone Support",
    "Reconstitution Supplies",
]

CAT_SLUGS = {
    "Metabolic & Weight Management":  "metabolic-weight-management",
    "Recovery & Regeneration":         "recovery-regeneration",
    "Growth Hormone & Secretagogues":  "growth-hormone-secretagogues",
    "Cognitive & Nootropics":          "cognitive-nootropics",
    "Copper Peptides & Aesthetic":     "copper-peptides-aesthetic",
    "Longevity & Bioregulators":       "longevity-bioregulators",
    "Hormone Support":                 "hormone-support",
    "Reconstitution Supplies":         "reconstitution-supplies",
}


def slugify(text):
    text = str(text).lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_-]+", "-", text)
    return text.strip("-")


def load_skus():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    skus = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        skus.append(dict(zip(headers, row)))
    return skus


def safe_float(val):
    try:
        return float(val) if val is not None else None
    except:
        return None


def main():
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found.")
        sys.exit(1)

    skus = load_skus()
    visible = [s for s in skus
               if s.get("Website Visible") == "Yes"
               and s.get("Status") == "Active"]

    # Build grouped structure
    categories = []
    for cat in CAT_ORDER:
        items = [s for s in visible if s.get("Category") == cat]
        if not items:
            continue

        compounds = []
        for s in items:
            compound = s.get("Compound", "")
            strength = s.get("Strength", "")
            fmt = s.get("Format", "")
            sku = s.get("SKU", "")
            retail = safe_float(s.get("Aura Retail Price"))
            tier = s.get("Tier", "")
            featured = s.get("Featured") == "Yes"

            # Build compound slug: compound-strength
            compound_slug = slugify(f"{compound}-{strength}")

            compounds.append({
                "sku":      sku,
                "compound": compound,
                "strength": strength,
                "format":   fmt,
                "price":    retail,
                "slug":     compound_slug,
                "tier":     tier,
                "featured": featured,
            })

        categories.append({
            "category":    cat,
            "slug":        CAT_SLUGS.get(cat, slugify(cat)),
            "compounds":   compounds,
        })

    # Full structured output
    output = {
        "_meta": {
            "source":    "MasterPricing.xlsx",
            "version":   "1.0",
            "generated": datetime.now().isoformat(),
            "effective": date.today().isoformat(),
            "standard":  "AURA LABS PRICING STANDARD v1.0",
            "count":     len(visible),
        },
        "categories": categories,
    }

    OUTPUT_JSON.write_text(json.dumps(output, indent=2), encoding="utf-8")
    print(f"✓ pricing.json: {len(visible)} SKUs across {len(categories)} categories")

    # Flat array for simple website lookups
    flat = []
    for cat_block in categories:
        for c in cat_block["compounds"]:
            flat.append({
                "category":      cat_block["category"],
                "category_slug": cat_block["slug"],
                **c
            })

    OUTPUT_FLAT.write_text(json.dumps(flat, indent=2), encoding="utf-8")
    print(f"✓ pricing-flat.json: flat array of {len(flat)} items")


if __name__ == "__main__":
    main()
