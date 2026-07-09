"""
Aura Labs Pricing Engine — Generator #3
Distributor Pricing PDF

Private document. Shows: Retail | Distributor Price (20% off) | MOQ

Input:  MasterPricing.xlsx
Output: output/Aura-Labs-Distributor-Pricing.pdf

Usage:  python generate-distributor-pricing.py
"""

import openpyxl
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path

SCRIPT_DIR  = Path(__file__).parent
XLSX        = SCRIPT_DIR / "MasterPricing.xlsx"
OUTPUT_HTML = SCRIPT_DIR / "output" / "_distributor.html"
OUTPUT_PDF  = SCRIPT_DIR / "output" / "Aura-Labs-Distributor-Pricing.pdf"

DISTRIBUTOR_DISCOUNT = 0.20   # 20% off retail
DEFAULT_MOQ = 10              # 10 units minimum

CAT_ORDER = [
    "Metabolic & Weight Management",
    "Recovery & Regeneration",
    "Growth Hormone & Secretagogues",
    "Cognitive & Nootropics",
    "Copper Peptides & Aesthetic",
    "Longevity & Bioregulators",
    "Hormone Support",
    "Reconstitution Supplies",
]


def load_skus():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    skus = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        skus.append(dict(zip(headers, row)))
    return skus


def fmt_price(val):
    if val is None:
        return "—"
    try:
        return f"${float(val):,.0f}"
    except:
        return str(val)


def dist_price(retail):
    if retail is None:
        return "—"
    try:
        p = float(retail) * (1 - DISTRIBUTOR_DISCOUNT)
        # Round to nearest dollar
        return f"${round(p):,}"
    except:
        return "—"


def main():
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found.")
        sys.exit(1)

    skus = load_skus()
    active = [s for s in skus if s.get("Status") == "Active"]

    grouped = {cat: [] for cat in CAT_ORDER}
    for s in active:
        cat = s.get("Category", "")
        if cat in grouped:
            grouped[cat].append(s)

    today = date.today().strftime("%B %Y")
    gen_ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    sections = ""
    for cat, items in grouped.items():
        if not items:
            continue
        rows = ""
        for item in items:
            compound = item.get("Compound", "")
            strength = item.get("Strength", "")
            fmt = item.get("Format", "")
            retail = item.get("Aura Retail Price")
            dp = dist_price(retail)
            blend_note = ""
            if "+" in str(strength):
                blend_note = f'<span class="blend-note">{strength}</span>'
                strength = ""
            rows += f"""
      <tr>
        <td class="compound-name">{compound}{blend_note}</td>
        <td>{strength}</td>
        <td class="format">{fmt}</td>
        <td>{fmt_price(retail)}</td>
        <td class="dist">{dp}</td>
        <td class="moq">{DEFAULT_MOQ}</td>
      </tr>"""

        sections += f"""
  <div class="section-title">{cat}</div>
  <table>
    <thead>
      <tr>
        <th>Compound</th><th>Strength</th><th>Format</th>
        <th style="text-align:right">Retail</th>
        <th style="text-align:right">Distributor</th>
        <th style="text-align:right">MOQ</th>
      </tr>
    </thead>
    <tbody>{rows}
    </tbody>
  </table>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Aura Labs — Distributor Pricing | {today}</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=EB+Garamond:wght@400;500;600&family=Inter:wght@400;500;600;700&display=swap');
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:'Inter',sans-serif; background:#fafaf8; color:#1a1a1a; font-size:10px; line-height:1.5; }}
  .page {{ max-width:900px; margin:0 auto; padding:48px 52px; }}
  .header {{ display:flex; justify-content:space-between; align-items:flex-end; border-bottom:1.5px solid #1a1a1a; padding-bottom:16px; margin-bottom:24px; }}
  .brand-name {{ font-family:'EB Garamond',serif; font-size:22px; font-weight:600; letter-spacing:0.08em; text-transform:uppercase; }}
  .brand-sub {{ font-size:8.5px; font-weight:500; letter-spacing:0.2em; text-transform:uppercase; color:#666; margin-top:3px; }}
  .header-meta {{ text-align:right; font-size:8.5px; color:#666; line-height:1.8; }}
  .header-meta strong {{ color:#1a1a1a; }}
  .confidential {{ background:#1a1a1a; color:#fff; padding:8px 14px; margin-bottom:16px; font-size:8px; letter-spacing:0.2em; text-transform:uppercase; text-align:center; }}
  .notice {{ background:#f0ede8; border-left:2px solid #1a1a1a; padding:10px 14px; margin-bottom:24px; font-size:8.5px; color:#444; line-height:1.6; }}
  .section-title {{ font-family:'EB Garamond',serif; font-size:13px; font-weight:600; letter-spacing:0.12em; text-transform:uppercase; border-bottom:0.5px solid #ccc; padding-bottom:6px; margin-bottom:10px; margin-top:22px; }}
  table {{ width:100%; border-collapse:collapse; margin-bottom:4px; }}
  thead tr {{ border-bottom:1px solid #1a1a1a; }}
  thead th {{ font-size:7.5px; font-weight:700; letter-spacing:0.15em; text-transform:uppercase; color:#666; padding:6px 8px 6px 0; text-align:left; }}
  thead th:last-child, thead th:nth-last-child(2), thead th:nth-last-child(3) {{ text-align:right; padding-right:0; }}
  tbody tr {{ border-bottom:0.5px solid #e8e5e0; }}
  tbody tr:last-child {{ border-bottom:none; }}
  tbody td {{ padding:7px 8px 7px 0; font-size:9.5px; vertical-align:middle; }}
  .dist {{ text-align:right; font-weight:700; color:#1a6b2a; font-size:10px; padding-right:0; }}
  .moq {{ text-align:right; color:#888; padding-right:0; }}
  td:nth-child(4) {{ text-align:right; }}
  .compound-name {{ font-weight:500; }}
  .format {{ color:#888; font-size:8.5px; }}
  .blend-note {{ font-size:8px; color:#888; display:block; }}
  .footer {{ margin-top:36px; border-top:0.5px solid #ccc; padding-top:14px; display:flex; justify-content:space-between; }}
  .footer p {{ font-size:7.5px; color:#888; line-height:1.8; }}
</style>
</head>
<body>
<div class="page">
  <div class="header">
    <div>
      <div class="brand-name">Aura Labs</div>
      <div class="brand-sub">Distributor Pricing — Confidential</div>
    </div>
    <div class="header-meta">
      <p><strong>Effective:</strong> {today}</p>
      <p><strong>Discount:</strong> {int(DISTRIBUTOR_DISCOUNT*100)}% off retail</p>
      <p>auralabsbio.co</p>
    </div>
  </div>
  <div class="confidential">Confidential — For Authorized Distributors Only — Do Not Distribute</div>
  <div class="notice">
    Distributor pricing reflects a {int(DISTRIBUTOR_DISCOUNT*100)}% discount off current retail pricing. MOQ applies per SKU per order. All compounds for research and laboratory use only. Not for human consumption.
  </div>
  {sections}
  <div class="footer">
    <div>
      <p>Aura Labs &nbsp;·&nbsp; auralabsbio.co &nbsp;·&nbsp; Distributor pricing valid for {today}</p>
      <p>Subject to change with 30-day notice &nbsp;·&nbsp; Confidential</p>
    </div>
    <div style="text-align:right">
      <p>Generated: {gen_ts}</p>
    </div>
  </div>
</div>
</body>
</html>"""

    OUTPUT_HTML.write_text(html, encoding="utf-8")
    result = subprocess.run(
        ["wkhtmltopdf", "--page-size", "Letter",
         "--margin-top", "0", "--margin-bottom", "0",
         "--margin-left", "0", "--margin-right", "0",
         "--enable-local-file-access",
         str(OUTPUT_HTML), str(OUTPUT_PDF)],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print("Error:", result.stderr[-300:])
        sys.exit(1)

    size_kb = OUTPUT_PDF.stat().st_size // 1024
    print(f"✓ Distributor Pricing: {OUTPUT_PDF.name} ({size_kb} KB)")


if __name__ == "__main__":
    main()
