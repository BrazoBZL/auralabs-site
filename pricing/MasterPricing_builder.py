"""
Aura Labs — MasterPricing.xlsx Builder
Creates the master workbook from the current known data.
Run once to bootstrap. After that, edit MasterPricing.xlsx directly.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

HEADERS = [
    "Category", "Tier", "Compound", "Strength", "Format",
    "SKU", "Supplier",
    "Supplier Cost",
    "Swiss Chems", "Biotech Peptides", "Limitless Life Nootropics",
    "Core Peptides", "Peptide Sciences",
    "Competitor Average",
    "Aura Retail Price",
    "Gross Margin", "Margin %",
    "Status", "Website Visible", "Featured", "Coming Soon",
    "Notes", "Last Updated"
]

# ── DATA ──────────────────────────────────────────────────────────────────────
# fmt: Category | Tier | Compound | Strength | Format | SKU | Supplier |
#      SupplierCost | SwissChems | BiotechPep | LimitlessLife | CorePep |
#      PepSci | CompAvg | AuraRetail | GM | GM% |
#      Status | WebVisible | Featured | ComingSoon | Notes | LastUpdated

SKUS = [
    # ── METABOLIC & WEIGHT MANAGEMENT ──────────────────────────────────────
    ["Metabolic & Weight Management","A","Semaglutide","10mg","Lyophilized Powder","SM10","SPL",3.50,None,85,90,140,100,104,89,85.50,"96.1%","Active","Yes","Yes","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Semaglutide","15mg","Lyophilized Powder","SM15","SPL",4.50,None,110,120,185,130,136,119,114.50,"96.2%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Semaglutide","20mg","Lyophilized Powder","SM20","SPL",5.50,None,140,150,230,165,171,149,143.50,"96.3%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Semaglutide","30mg","Lyophilized Powder","SM30","SPL",7.50,None,200,210,320,240,242,199,191.50,"96.2%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Tirzepatide","10mg","Lyophilized Powder","TR10","SPL",5.50,90,95,90,130,110,103,89,83.50,"95.4%","Active","Yes","Yes","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Tirzepatide","15mg","Lyophilized Powder","TR15","SPL",7.50,120,130,120,175,145,138,119,111.50,"93.7%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Tirzepatide","20mg","Lyophilized Powder","TR20","SPL",9.50,150,165,150,215,175,171,149,139.50,"93.6%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Tirzepatide","30mg","Lyophilized Powder","TR30","SPL",11.50,200,210,200,280,230,224,189,177.50,"93.9%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Tirzepatide","50mg","Lyophilized Powder","TR50","SPL",15.30,310,330,315,430,365,350,299,283.70,"94.9%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Tirzepatide","60mg","Lyophilized Powder","TR60","SPL",17.50,355,370,350,490,410,395,329,311.50,"94.7%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Retatrutide","10mg","Lyophilized Powder","RT10","SPL",10.50,95,110,110,145,135,119,99,88.50,"89.4%","Active","Yes","Yes","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Retatrutide","15mg","Lyophilized Powder","RT15","SPL",12.50,130,145,150,195,175,159,139,126.50,"91.0%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Retatrutide","20mg","Lyophilized Powder","RT20","SPL",14.50,165,185,190,245,220,201,169,154.50,"91.4%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Retatrutide","30mg","Lyophilized Powder","RT30","SPL",16.50,215,240,245,315,285,260,219,202.50,"92.5%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Retatrutide","50mg","Lyophilized Powder","RT50","SPL",25.00,345,380,380,490,450,409,349,324.00,"92.8%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","Retatrutide","60mg","Lyophilized Powder","RT60","SPL",27.50,395,430,430,555,510,464,389,361.50,"92.9%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","A","AOD-9604","5mg","Lyophilized Powder","AOD5","SPL",10.00,60,70,75,73,68,69,59,49.00,"83.1%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","B","L-Carnitine","600mg / 10ml","Injectable Solution","LC600","SPL",8.00,None,65,68,65,60,63,59,51.00,"86.4%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","B","Lemon Bottle","10ml","Injectable Solution","LB10","SPL",8.00,110,115,120,115,110,114,99,91.00,"91.9%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","B","Lipo-C","10ml","Injectable Solution","LC120","SPL",8.00,None,95,99,95,88,93,79,71.00,"89.9%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","B","Lipo-C FatBlaster","10ml","Injectable Solution","LC526","SPL",8.00,None,100,99,100,92,97,89,81.00,"91.0%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","B","Lipo-C with B12","10ml","Injectable Solution","LC216","SPL",8.00,None,100,99,100,92,97,89,81.00,"91.0%","Active","Yes","No","No","","2026-07-09"],
    ["Metabolic & Weight Management","B","Super Human Blend","10ml","Injectable Solution","SHB","SPL",8.00,100,110,115,110,105,108,99,91.00,"91.9%","Active","Yes","No","No","","2026-07-09"],
    # ── RECOVERY & REGENERATION ──────────────────────────────────────────────
    ["Recovery & Regeneration","A","BPC-157","5mg","Lyophilized Powder","BC5","SPL",4.80,36,47,50,52,45,46,39,34.20,"87.7%","Active","Yes","Yes","No","","2026-07-09"],
    ["Recovery & Regeneration","A","BPC-157","10mg","Lyophilized Powder","BC10","SPL",6.50,70,85,100,97,75,85,69,62.50,"90.6%","Active","Yes","No","No","","2026-07-09"],
    ["Recovery & Regeneration","A","BPC-157","20mg","Lyophilized Powder","BC20","SPL",9.20,120,145,165,175,135,148,119,109.80,"92.3%","Active","Yes","No","No","","2026-07-09"],
    ["Recovery & Regeneration","A","TB-500 (Full-Chain)","5mg","Lyophilized Powder","TB5","SPL",12.50,39,72,55,78,60,61,59,46.50,"78.8%","Active","Yes","Yes","No","Full-chain B4 Acetate superior grade","2026-07-09"],
    ["Recovery & Regeneration","A","TB-500 (Full-Chain)","10mg","Lyophilized Powder","TB10","SPL",18.50,75,124,134,140,100,115,99,80.50,"81.3%","Active","Yes","No","No","Full-chain B4 Acetate superior grade","2026-07-09"],
    ["Recovery & Regeneration","B","KPV","5mg","Lyophilized Powder","KPV5","SPL",5.50,50,50,49,55,52,51,49,43.50,"88.8%","Active","Yes","No","No","","2026-07-09"],
    ["Recovery & Regeneration","B","KPV","10mg","Lyophilized Powder","KPV10","SPL",6.80,80,80,79,90,78,81,69,62.20,"90.1%","Active","Yes","No","No","","2026-07-09"],
    ["Recovery & Regeneration","B","MOTS-C","10mg","Lyophilized Powder","MS10","SPL",6.60,65,70,85,75,80,75,69,62.40,"90.4%","Active","Yes","No","No","","2026-07-09"],
    ["Recovery & Regeneration","B","MOTS-C","20mg","Lyophilized Powder","MS20","SPL",10.60,110,115,140,125,130,124,109,98.40,"90.3%","Active","Yes","No","No","","2026-07-09"],
    ["Recovery & Regeneration","B","MOTS-C","40mg","Lyophilized Powder","MS40","SPL",16.50,170,175,200,185,195,185,169,152.50,"90.2%","Active","Yes","No","No","","2026-07-09"],
    ["Recovery & Regeneration","A","BB10 — BPC-157 + TB-500","5mg + 5mg","Lyophilized Powder","BB10","SPL",13.00,None,110,115,105,108,108,89,76.00,"85.4%","Active","Yes","No","No","","2026-07-09"],
    ["Recovery & Regeneration","A","BB20 — BPC-157 + TB-500","10mg + 10mg","Lyophilized Powder","BB20","SPL",21.80,None,180,195,195,175,181,149,127.20,"85.4%","Active","Yes","No","No","","2026-07-09"],
    ["Recovery & Regeneration","A","GLOW70 — BPC-157 + TB-500 + GHK-Cu","10mg+10mg+50mg","Lyophilized Powder","GLOW70","SPL",25.00,None,270,280,300,265,272,229,204.00,"89.1%","Active","Yes","No","No","","2026-07-09"],
    ["Recovery & Regeneration","A","KLOW80 — KPV + BPC + TB-500 + GHK-Cu","10mg+10mg+10mg+50mg","Lyophilized Powder","KLOW80","SPL",29.00,None,315,325,350,305,317,269,240.00,"89.2%","Active","Yes","No","No","","2026-07-09"],
    # ── GROWTH HORMONE & SECRETAGOGUES ──────────────────────────────────────
    ["Growth Hormone & Secretagogues","B","CJC-1295 (no DAC)","10mg","Lyophilized Powder","CJC10","SPL",13.60,86,95,100,92,88,92,79,65.40,"82.8%","Active","Yes","No","No","","2026-07-09"],
    ["Growth Hormone & Secretagogues","B","CJC-1295 + Ipamorelin","5mg + 5mg","Lyophilized Powder","CP10","SPL",13.50,80,88,95,92,85,88,79,65.50,"82.9%","Active","Yes","No","No","Blend","2026-07-09"],
    ["Growth Hormone & Secretagogues","B","Tesamorelin","5mg","Lyophilized Powder","TSM5","SPL",9.50,55,60,65,65,70,63,59,49.50,"83.9%","Active","Yes","No","No","","2026-07-09"],
    ["Growth Hormone & Secretagogues","B","Tesamorelin","10mg","Lyophilized Powder","TSM10","SPL",14.50,100,120,125,130,135,122,109,94.50,"86.7%","Active","Yes","No","No","","2026-07-09"],
    # ── COGNITIVE & NOOTROPICS ───────────────────────────────────────────────
    ["Cognitive & Nootropics","B","NAD+","100mg","Lyophilized Powder","NJ100","SPL",4.60,48,40,45,47,40,44,39,34.40,"88.2%","Active","Yes","Yes","No","","2026-07-09"],
    ["Cognitive & Nootropics","B","NAD+","500mg","Lyophilized Powder","NJ500","SPL",7.50,110,80,90,87,94,92,79,71.50,"90.5%","Active","Yes","No","No","","2026-07-09"],
    ["Cognitive & Nootropics","B","NAD+","1000mg","Lyophilized Powder","NJ1000","SPL",10.00,165,130,145,140,130,142,129,119.00,"92.2%","Active","Yes","No","No","","2026-07-09"],
    ["Cognitive & Nootropics","B","Semax","10mg","Lyophilized Powder","XA10","SPL",4.60,45,45,60,55,50,51,49,44.40,"90.6%","Active","Yes","No","No","","2026-07-09"],
    ["Cognitive & Nootropics","B","Semax","30mg","Lyophilized Powder","XA30","SPL",11.00,42,65,77,80,65,66,59,48.00,"81.4%","Active","Yes","No","No","","2026-07-09"],
    ["Cognitive & Nootropics","B","Selank","10mg","Lyophilized Powder","SK10","SPL",4.60,50,48,52,55,50,51,49,44.40,"90.6%","Active","Yes","No","No","","2026-07-09"],
    ["Cognitive & Nootropics","B","Selank","30mg","Lyophilized Powder","SK30","SPL",11.00,80,95,89,95,90,90,79,68.00,"86.1%","Active","Yes","No","No","","2026-07-09"],
    ["Cognitive & Nootropics","B","Glutathione","600mg","Lyophilized Powder","GT600","SPL",5.00,55,65,72,68,62,64,59,54.00,"91.5%","Active","Yes","No","No","","2026-07-09"],
    # ── COPPER PEPTIDES & AESTHETIC ──────────────────────────────────────────
    ["Copper Peptides & Aesthetic","B","GHK-Cu","50mg","Lyophilized Powder","CU50","SPL",4.90,57,45,79,58,45,57,49,44.10,"90.0%","Active","Yes","Yes","No","","2026-07-09"],
    ["Copper Peptides & Aesthetic","B","GHK-Cu","100mg","Lyophilized Powder","CU100","SPL",6.30,88,70,90,90,68,81,69,62.70,"90.9%","Active","Yes","No","No","","2026-07-09"],
    ["Copper Peptides & Aesthetic","B","Healthy Hair Skin Nails Blend","10ml","Injectable Solution","HHB","SPL",8.00,None,105,110,108,99,108,89,81.00,"91.0%","Active","Yes","No","No","","2026-07-09"],
    # ── LONGEVITY & BIOREGULATORS (coming soon) ──────────────────────────────
    ["Longevity & Bioregulators","C","FOXO4-DRI","2mg","Lyophilized Powder","FX2","TBD",None,None,None,None,None,None,None,None,None,"N/A","Coming Soon","No","No","Yes","Sourcing required","2026-07-09"],
    ["Longevity & Bioregulators","C","ACE-031","1mg","Lyophilized Powder","AC1","TBD",None,None,None,None,None,None,None,None,None,"N/A","Coming Soon","No","No","Yes","Sourcing required","2026-07-09"],
    ["Longevity & Bioregulators","C","ARA-290","16mg","Lyophilized Powder","AR16","TBD",None,None,None,None,None,None,None,None,None,"N/A","Coming Soon","No","No","Yes","Sourcing required","2026-07-09"],
    # ── HORMONE SUPPORT (coming soon) ────────────────────────────────────────
    ["Hormone Support","C","HGH","IU","Lyophilized Powder","HGH","TBD",None,None,None,None,None,None,None,None,None,"N/A","Coming Soon","No","No","Yes","Sourcing required","2026-07-09"],
    ["Hormone Support","C","HCG","IU","Lyophilized Powder","HCG","TBD",None,None,None,None,None,None,None,None,None,"N/A","Coming Soon","No","No","Yes","Sourcing required","2026-07-09"],
    # ── RECONSTITUTION SUPPLIES ──────────────────────────────────────────────
    ["Reconstitution Supplies","B","Bacteriostatic Water","3ml","Sterile Solution","WA3","SPL",0.80,None,None,None,None,None,None,9,8.20,"91.1%","Active","Yes","No","No","","2026-07-09"],
    ["Reconstitution Supplies","B","Bacteriostatic Water","5ml","Sterile Solution","WA5","SPL",1.00,None,None,None,None,None,None,12,11.00,"91.7%","Active","Yes","No","No","","2026-07-09"],
    ["Reconstitution Supplies","B","Bacteriostatic Water","10ml","Sterile Solution","WA10","SPL",1.00,None,None,None,None,None,None,15,14.00,"93.3%","Active","Yes","No","No","","2026-07-09"],
    ["Reconstitution Supplies","B","Benzyl Alcohol 0.9%","3ml","Sterile Solution","BA3","SPL",0.80,None,None,None,None,None,None,9,8.20,"91.1%","Active","Yes","No","No","","2026-07-09"],
    ["Reconstitution Supplies","B","Benzyl Alcohol 0.9%","5ml","Sterile Solution","BA5","SPL",1.00,None,None,None,None,None,None,12,11.00,"91.7%","Active","Yes","No","No","","2026-07-09"],
    ["Reconstitution Supplies","B","Benzyl Alcohol 0.9%","10ml","Sterile Solution","BA10","SPL",1.00,None,None,None,None,None,None,15,14.00,"93.3%","Active","Yes","No","No","","2026-07-09"],
    ["Reconstitution Supplies","B","Acetic Acid 0.6%","3ml","Sterile Solution","AA3","SPL",0.80,None,None,None,None,None,None,9,8.20,"91.1%","Active","Yes","No","No","","2026-07-09"],
    ["Reconstitution Supplies","B","Acetic Acid 0.6%","10ml","Sterile Solution","AA10","SPL",1.00,None,None,None,None,None,None,15,14.00,"93.3%","Active","Yes","No","No","","2026-07-09"],
]

# ── STYLING ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
CAT_FILL    = PatternFill("solid", fgColor="F0EDE8")
CAT_FONT    = Font(bold=True, name="Calibri", size=10)

TIER_COLORS = {"A": "E8F4E8", "B": "E8EEF4", "C": "F4E8F0"}

def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Pricing"

    # Header row
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28

    # Data rows
    current_cat = None
    row = 2
    for sku in SKUS:
        cat = sku[0]
        if cat != current_cat:
            current_cat = cat

        for col, val in enumerate(sku, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(vertical="center")
            # Tier color
            tier = sku[1]
            if tier in TIER_COLORS:
                cell.fill = PatternFill("solid", fgColor=TIER_COLORS[tier])
            # Price formatting
            if col in (8,9,10,11,12,13,14,15,16) and val is not None:
                cell.number_format = '"$"#,##0.00'

        row += 1

    # Column widths
    widths = {
        1: 28, 2: 6, 3: 36, 4: 14, 5: 20, 6: 10,
        7: 8, 8: 12, 9: 12, 10: 14, 11: 18, 12: 12,
        13: 14, 14: 16, 15: 14, 16: 14, 17: 10,
        18: 10, 19: 12, 20: 10, 21: 12,
        22: 30, 23: 14
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    # Metadata sheet
    meta = wb.create_sheet("Metadata")
    meta["A1"] = "Aura Labs Pricing Engine"
    meta["A2"] = "Version"
    meta["B2"] = "1.0"
    meta["A3"] = "Pricing Standard"
    meta["B3"] = "AURA LABS PRICING STANDARD v1.0"
    meta["A4"] = "Supplier"
    meta["B4"] = "Science Peptide Lab (SPL)"
    meta["A5"] = "Benchmarks"
    meta["B5"] = "Swiss Chems | Biotech Peptides | Limitless Life Nootropics | Core Peptides | Peptide Sciences"
    meta["A6"] = "Last Rebuilt"
    meta["B6"] = str(date.today())
    meta["A7"] = "Tier A Target"
    meta["B7"] = "~15% below benchmark average"
    meta["A8"] = "Tier B Target"
    meta["B8"] = "~10% below benchmark average"
    meta["A9"] = "Tier C Target"
    meta["B9"] = "Market average to 5% below"

    for cell in ["A1","A2","A3","A4","A5","A6","A7","A8","A9"]:
        meta[cell].font = Font(bold=True)

    output = "MasterPricing.xlsx"
    wb.save(output)
    print(f"✓ Built {output} — {len(SKUS)} SKUs")

if __name__ == "__main__":
    build()
